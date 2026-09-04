"""
tests/test_core_logic.py
Model-free tests of the correctness-critical logic:
  - tracker persistence + occlusion recovery
  - identity evidence state machine (no single-frame identity)
  - attendance fusion (one record per employee across videos)
  - DB UNIQUE(employee_id, attendance_date) + UPSERT
  - Excel export uniqueness

Run from the project folder:  python -m pytest tests -q
"""
from __future__ import annotations
import os
import sys
from datetime import datetime

import numpy as np
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import settings                                    # noqa: E402
from core.byte_track import ByteTracker                        # noqa: E402
from core import identity_evidence as ie                       # noqa: E402
from core.attendance_engine import (TrackEvidence, fuse_day,   # noqa: E402
                                    full_roster_view, PRESENT)
from core import reid                                          # noqa: E402


def _box(x, y, w=60, h=140):
    return (x, y, x + w, y + h)


# ── tracker ────────────────────────────────────────────────────────────────
def test_track_id_persists_while_walking():
    tr = ByteTracker()
    tid = None
    for i in range(8):
        out = tr.update([(_box(100 + i * 12, 200), 0.9)])
        if out:
            ids = {t.track_id for t in out}
            assert len(ids) == 1
            tid = tid or next(iter(ids))
            assert next(iter(ids)) == tid


def test_track_survives_short_occlusion():
    tr = ByteTracker()
    for i in range(4):
        out = tr.update([(_box(100 + i * 10, 200), 0.9)])
    tid = out[0].track_id
    for _ in range(3):                       # person occluded: no detections
        tr.update([])
    out = tr.update([(_box(170, 200), 0.9)])  # reappears near prediction
    assert out and out[0].track_id == tid


def test_low_confidence_detection_keeps_track_alive():
    tr = ByteTracker()
    for i in range(4):
        out = tr.update([(_box(100, 200), 0.9)])
    tid = out[0].track_id
    out = tr.update([(_box(104, 202), 0.3)])   # partially occluded, low score
    assert out and out[0].track_id == tid


def test_two_people_two_tracks():
    tr = ByteTracker()
    for _ in range(4):
        out = tr.update([(_box(100, 200), 0.9), (_box(500, 200), 0.9)])
    assert len({t.track_id for t in out}) == 2


# ── identity evidence: never trust one frame ───────────────────────────────
def _obs(emp, sim=0.6, strong=True, accepted=True, ambiguous=False):
    return ie.Observation(emp, sim, 0.7 if strong else 0.45,
                          strong, accepted, ambiguous)


def test_single_match_does_not_confirm():
    t = ie.TrackIdentity(track_id=1)
    ie.update(t, _obs("E001"))
    assert t.status in (ie.CANDIDATE, ie.UNKNOWN)
    assert t.status not in (ie.CONFIRMED, ie.LOCKED)


def test_accumulated_evidence_confirms_and_locks():
    t = ie.TrackIdentity(track_id=1)
    for _ in range(settings.IDENTITY_LOCK_THRESHOLD):
        ie.update(t, _obs("E001", sim=0.62))
    assert t.status == ie.LOCKED and t.identity == "E001"
    assert not t.needs_recognition()          # locked track stops paying CPU


def test_conflicting_strong_evidence_goes_to_review():
    t = ie.TrackIdentity(track_id=1)
    for _ in range(settings.MIN_STRONG_VOTES):
        ie.update(t, _obs("E001"))
    for _ in range(settings.MIN_STRONG_VOTES):
        ie.update(t, _obs("E002"))
    assert t.status == ie.REVIEW_REQUIRED


# ── attendance fusion ──────────────────────────────────────────────────────
def _ev(emp, video, conf=0.8, status=ie.CONFIRMED, n=6):
    now = datetime(2026, 7, 10, 9, 0, 0)
    return TrackEvidence(employee_id=emp, status=status, confidence=conf,
                         first_seen=now, last_seen=now, camera_id="C1",
                         camera_location="office", camera_type="office_room",
                         video_name=video, n_accepted=n)


def test_same_employee_many_tracks_many_videos_one_record():
    evs = ([_ev("E004", "v1.mp4")] * 3 + [_ev("E004", "v2.mp4")] * 2
           + [_ev("E001", "v1.mp4")])
    recs = fuse_day("2026-07-10", evs)
    assert set(recs) == {"E004", "E001"}          # dict keyed by id: one each
    assert recs["E004"].n_tracks == 5
    assert sorted(recs["E004"].videos_seen) == ["v1.mp4", "v2.mp4"]
    assert recs["E004"].attendance_status == PRESENT


def test_multi_video_present_set_merges():
    evs = [_ev("E001", "v1"), _ev("E004", "v1"), _ev("E009", "v1"),
           _ev("E002", "v2"), _ev("E004", "v2"), _ev("E010", "v2"),
           _ev("E003", "v3"), _ev("E009", "v3")]
    recs = fuse_day("2026-07-10", evs)
    assert set(recs) == {"E001", "E002", "E003", "E004", "E009", "E010"}


def test_unconfirmed_or_weak_tracks_do_not_mark_present():
    recs = fuse_day("2026-07-10", [
        _ev("E007", "v1", status=ie.CANDIDATE),      # sustained candidate:
        _ev("E008", "v1", conf=0.40),                # surfaces as REVIEW only
    ])
    assert recs["E007"].attendance_status == "REVIEW"       # visible, but…
    assert recs["E007"].review_required
    assert recs["E007"].attendance_status != PRESENT        # …never PRESENT
    assert recs["E008"].attendance_status != PRESENT
    assert recs["E008"].review_required


def test_full_roster_marks_absent_once_each():
    recs = fuse_day("2026-07-10", [_ev("E001", "v1")])
    rows = full_roster_view("2026-07-10", recs, ["E001", "E002", "E003"])
    assert [r["employee_id"] for r in rows] == ["E001", "E002", "E003"]
    assert rows[1]["status"] == "ABSENT" and rows[2]["status"] == "ABSENT"


# ── database uniqueness ────────────────────────────────────────────────────
def test_db_upsert_never_duplicates(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "DB_PATH", str(tmp_path / "t.db"))
    from database import db_setup, db_manager
    db_setup.init_db()
    db_manager.add_employee("E001", "Asha", "Ops")
    rec = fuse_day("2026-07-10", [_ev("E001", "v1.mp4")])["E001"]
    for _ in range(3):                              # save repeatedly (re-runs)
        db_manager.save_daily_record(rec)
    rows = db_manager.get_daily_attendance("2026-07-10")
    assert len(rows) == 1 and rows[0]["status"] == PRESENT

    import sqlite3
    con = sqlite3.connect(settings.DB_PATH)
    n = con.execute("SELECT COUNT(*) FROM attendance").fetchone()[0]
    con.close()
    assert n == 1                                    # DB-level guarantee


def test_db_raw_duplicate_insert_is_impossible(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "DB_PATH", str(tmp_path / "t2.db"))
    from database import db_setup
    import sqlite3
    db_setup.init_db()
    con = sqlite3.connect(settings.DB_PATH)
    con.execute("INSERT INTO employees(employee_id, name) VALUES('E1','A')")
    con.execute("INSERT INTO attendance(employee_id, attendance_date, status) "
                "VALUES ('E1','2026-07-10','PRESENT')")
    with pytest.raises(sqlite3.IntegrityError):
        con.execute("INSERT INTO attendance(employee_id, attendance_date, status) "
                    "VALUES ('E1','2026-07-10','PRESENT')")
    con.close()


# ── excel export ───────────────────────────────────────────────────────────
def test_excel_one_row_per_employee(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "EXPORTS_DIR", str(tmp_path))
    from reports import excel_report
    rows = [{"employee_id": "E001", "name": "Asha", "department": "Ops",
             "attendance_date": "2026-07-10", "status": "PRESENT",
             "confidence": 0.82, "evidence_type": "face",
             "face_evidence_count": 9, "body_evidence_count": 0,
             "videos_seen": "v1.mp4", "review_required": 0}] * 3   # dup input
    path = excel_report.export("2026-07-10", rows)
    from openpyxl import load_workbook
    ws = load_workbook(path)["Attendance"]
    data_rows = [r for r in ws.iter_rows(min_row=3, values_only=True)
                 if r[0] == "E001"]
    assert len(data_rows) == 1


# ── reid memory (support evidence only) ────────────────────────────────────
def test_reid_descriptor_and_linking():
    rng = np.random.default_rng(0)
    crop_a = (rng.random((120, 50, 3)) * 255).astype("uint8")
    d1 = reid.appearance_descriptor(crop_a)
    d2 = reid.appearance_descriptor(crop_a.copy())
    assert d1 is not None and float(np.dot(d1, d2)) > 0.99
    mem = reid.ReIDMemory()
    mem.observe(7, d1, sample_idx=0)
    link, sim = mem.best_link(d2, exclude=99)
    assert link == 7 and sim > 0.99
    assert reid.appearance_descriptor(np.zeros((10, 5, 3), "uint8")) is None


# ── recognition focus scheduler: one person at a time ──────────────────────
from core.byte_track import Track
from core.recognition_scheduler import RecognitionScheduler


def _trk(tid, area_side=100):
    return Track(track_id=tid, box=(0, 0, area_side, area_side * 2), score=0.9)


class _Ident:
    def __init__(self, needs=True, n_accepted=0):
        self._needs = needs
        self.n_accepted = n_accepted

    def needs_recognition(self):
        return self._needs


def test_scheduler_focuses_exactly_one_person():
    s = RecognitionScheduler(focus_limit=1, patience=5, cooldown=10)
    tracks = [_trk(1, 80), _trk(2, 120), _trk(3, 60)]
    idents = {1: _Ident(), 2: _Ident(), 3: _Ident()}
    picked = s.pick(tracks, idents, sample_idx=0)
    assert len(picked) == 1
    assert picked == {2}                      # largest box (closest) first


def test_scheduler_prefers_person_already_in_progress():
    s = RecognitionScheduler(focus_limit=1, patience=5, cooldown=10)
    tracks = [_trk(1, 200), _trk(2, 80)]
    idents = {1: _Ident(), 2: _Ident(n_accepted=3)}   # 2 has votes already
    assert s.pick(tracks, idents, 0) == {2}   # finish them despite smaller box


def test_scheduler_sticks_until_resolved_then_moves_on():
    s = RecognitionScheduler(focus_limit=1, patience=5, cooldown=10)
    tracks = [_trk(1, 120), _trk(2, 80)]
    idents = {1: _Ident(), 2: _Ident()}
    assert s.pick(tracks, idents, 0) == {1}
    s.report(1, 0, progressed=True, resolved=False)
    assert s.pick(tracks, idents, 1) == {1}   # still working on person 1
    s.report(1, 1, progressed=True, resolved=True)   # confirmed!
    idents[1] = _Ident(needs=False)
    assert s.pick(tracks, idents, 2) == {2}   # next person


def test_scheduler_rotates_when_face_never_appears():
    s = RecognitionScheduler(focus_limit=1, patience=3, cooldown=50)
    tracks = [_trk(1, 120), _trk(2, 80)]
    idents = {1: _Ident(), 2: _Ident()}
    assert s.pick(tracks, idents, 0) == {1}
    for i in range(3):                        # person 1 keeps facing away
        s.report(1, i, progressed=False, resolved=False)
    assert s.pick(tracks, idents, 3) == {2}   # patience exhausted -> next person
    # and person 1 is retried once nobody else is waiting
    idents[2] = _Ident(needs=False)
    assert s.pick(tracks, idents, 4) == {1}


def test_track_survives_box_size_flicker():
    """Seated person: box flips from full-body to upper-body (IoU < match
    threshold) — the centre-distance rescue must keep the SAME track id."""
    tr = ByteTracker(match_iou=0.30)
    out = None
    for _ in range(4):                                    # establish track
        out = tr.update([((100, 100, 160, 240), 0.9)])
    tid = out[0].track_id
    # upper-body-only box: IoU vs (100,100,160,240) ≈ 0.21 -> IoU match fails
    out = tr.update([((110, 105, 150, 150), 0.9)])
    assert out and out[0].track_id == tid
    assert len(tr.all_tracks()) == 1                      # no fragment created


def test_scheduler_prefers_visible_face_over_big_box():
    """Ceiling-camera reality: nearest person (huge box) faces AWAY; a
    mid-distance person shows a clear face. Focus must go to the face."""
    s = RecognitionScheduler(focus_limit=1, patience=10, cooldown=20)
    tracks = [_trk(1, 300), _trk(2, 120)]        # 1 = huge box, no face
    idents = {1: _Ident(), 2: _Ident()}
    faces = {2: 48.0}                             # only person 2 shows a face
    assert s.pick(tracks, idents, 0, faces) == {2}


def test_scheduler_preempts_faceless_focus():
    """Focused person turns away while someone else shows a face -> switch
    immediately instead of burning patience."""
    s = RecognitionScheduler(focus_limit=1, patience=10, cooldown=20)
    tracks = [_trk(1, 300), _trk(2, 120)]
    idents = {1: _Ident(), 2: _Ident()}
    assert s.pick(tracks, idents, 0, {1: 60.0}) == {1}     # 1 has the face
    # next sample: 1's face gone, 2's face appears -> preempt to 2
    assert s.pick(tracks, idents, 1, {2: 50.0}) == {2}


# ── false-identity protections (10 Jul incident) ───────────────────────────
def _confirmed_track(emp="02", n=8):
    t = ie.TrackIdentity(track_id=99)
    for _ in range(n):
        ie.update(t, _obs(emp, sim=0.62))
    assert t.status in (ie.CONFIRMED, ie.LOCKED)
    return t


def test_inherited_evidence_can_never_confirm_alone():
    """A stranger appearance-linked to a lost confirmed track must NOT be
    confirmed from inherited (clothing-based) evidence."""
    donor = _confirmed_track()
    stranger = ie.TrackIdentity(track_id=7)
    ie.inherit(stranger, donor)
    assert stranger.status not in (ie.CONFIRMED, ie.LOCKED)
    assert stranger.needs_recognition()


def test_inheritance_plus_fresh_face_evidence_confirms():
    """The rightful person's continuation confirms quickly: head start +
    a couple of FRESH strong face matches."""
    donor = _confirmed_track()
    cont = ie.TrackIdentity(track_id=8)
    ie.inherit(cont, donor)
    for _ in range(2):
        ie.update(cont, _obs("02", sim=0.60))
    assert cont.status in (ie.CONFIRMED, ie.LOCKED)


def test_weaken_strips_conflicting_identity():
    t = _confirmed_track()
    ie.weaken(t, "02")
    assert t.status not in (ie.CONFIRMED, ie.LOCKED)
    assert t.needs_recognition()


# ── Phase 2: body ReID is SUPPORT-only ──────────────────────────────────────
def test_body_support_alone_can_never_confirm():
    t = ie.TrackIdentity(track_id=1)
    for _ in range(200):                      # body spam for hours
        ie.support(t, "02", 0.75)
    assert t.status not in (ie.CONFIRMED, ie.LOCKED)
    assert t.needs_recognition()
    assert t.votes["02"] <= settings.BODY_SUPPORT_CAP + 1e-6   # capped


def test_body_support_cannot_flip_a_face_confirmed_identity():
    t = _confirmed_track("01", n=8)           # face-proven as 01
    for _ in range(500):                      # adversarial body spam for 02
        ie.support(t, "02", 0.80)
    assert t.identity == "01"                 # face outweighs capped body votes
    assert t.status in (ie.CONFIRMED, ie.LOCKED)


def test_body_support_speeds_confirmation_of_right_person():
    a = ie.TrackIdentity(track_id=1)          # with body support
    b = ie.TrackIdentity(track_id=2)          # without
    for _ in range(6):
        ie.support(a, "02", 0.7)
    for t in (a, b):
        for _ in range(settings.MIN_IDENTITY_VOTES):
            ie.update(t, _obs("02", sim=0.55))
    assert a.status in (ie.CONFIRMED, ie.LOCKED)
    assert a.vote_ratio >= b.vote_ratio       # support strengthened, not replaced


def test_body_gallery_match_and_dedup(tmp_path, monkeypatch):
    import core.body_gallery as bgal
    # body templates are day-scoped since 31 Jul 2026 (clothes are only valid
    # for the day they were harvested), so the root is what gets redirected.
    monkeypatch.setattr(bgal, "BODY_EMB_ROOT", str(tmp_path))
    g = bgal.BodyGallery(day="2026-07-31")
    v1 = np.zeros(settings.BODY_EMBED_DIM, dtype=np.float32); v1[0] = 1.0
    v2 = np.zeros(settings.BODY_EMBED_DIM, dtype=np.float32); v2[1] = 1.0
    assert g.add_embedding("01", v1)
    assert not g.add_embedding("01", v1)      # duplicate rejected
    assert g.add_embedding("02", v2)
    m = g.match(v1)
    assert m.employee_id == "01" and m.supported
    probe = np.zeros(settings.BODY_EMBED_DIM, dtype=np.float32)
    probe[0] = probe[1] = 0.7071              # equidistant -> margin too small
    assert not g.match(probe).supported


# ── Phase 2d: small-face upscale probe ──────────────────────────────────────
def test_probe_upscales_small_roi_and_returns_src_coords():
    from core.pipeline import probe_faces
    from models.face_detector import Face

    class StubDet:
        def __init__(self):
            self.calls = []
        def detect(self, img, offset=(0, 0)):
            self.calls.append(img.shape)
            if img.shape[0] < 200:            # native small ROI: nothing found
                return []
            return [Face(box=(10, 12, 58, 70),  # found on the upscaled copy
                         landmarks=np.zeros((5, 2), np.float32), score=0.8)]

    frame = np.zeros((1080, 1920, 3), np.uint8)
    det = StubDet()
    out = probe_faces(det, frame, (100, 100, 180, 260))   # 80x160 person ROI
    assert out is not None
    f, src, upscaled = out
    assert upscaled is True
    assert src.shape[0] == 320 and src.shape[1] == 160    # 2x super-sampled
    assert f.box == (10, 12, 58, 70)                      # src coordinates
    assert len(det.calls) == 2                            # native try, then 2x


def test_probe_skips_upscale_for_large_roi():
    from core.pipeline import probe_faces

    class NoneDet:
        def __init__(self): self.calls = 0
        def detect(self, img, offset=(0, 0)):
            self.calls += 1; return []

    frame = np.zeros((1080, 1920, 3), np.uint8)
    det = NoneDet()
    assert probe_faces(det, frame, (0, 0, 300, 700)) is None
    assert det.calls == 1                     # big ROI: no upscale retry


# ── 11 Jul false-ID fix: borderline sims vote, only >=0.50 confirms ────────
def test_borderline_similarity_can_never_confirm():
    """A lookalike stranger matching at 0.40-0.47 (quality fine) accumulates
    votes but must NEVER be confirmed."""
    t = ie.TrackIdentity(track_id=1)
    for _ in range(12):
        ie.update(t, ie.Observation("02", 0.44, 0.75, True, True, False))
    assert t.status not in (ie.CONFIRMED, ie.LOCKED)


def test_confirmation_grade_similarity_still_confirms():
    t = ie.TrackIdentity(track_id=1)
    for _ in range(settings.MIN_IDENTITY_VOTES):
        ie.update(t, ie.Observation("02", 0.56, 0.75, True, True, False))
    assert t.status in (ie.CONFIRMED, ie.LOCKED)


def test_sustained_candidate_surfaces_as_review_never_present():
    evs = [_ev("01", "v1", conf=0.95, status=ie.CANDIDATE, n=6)] * 3
    recs = fuse_day("2026-07-11", evs)
    assert "01" in recs
    assert recs["01"].attendance_status == "REVIEW"
    assert recs["01"].review_required
    assert recs["01"].attendance_status != PRESENT


def test_weak_candidate_below_review_floor_is_ignored():
    recs = fuse_day("2026-07-11",
                    [_ev("01", "v1", status=ie.CANDIDATE, n=1)])
    assert "01" not in recs


def test_confirmed_evidence_still_marks_present():
    recs = fuse_day("2026-07-11", [_ev("02", "v1", conf=0.85)])
    assert recs["02"].attendance_status == PRESENT


# ── Phase 3a: throughput machinery ──────────────────────────────────────────
def test_tracker_coast_returns_confirmed_without_aging():
    tr = ByteTracker()
    for i in range(4):
        tr.update([(_box(100 + i * 5, 200), 0.9)])
    before = tr.coast()
    assert len(before) == 1
    missed_before = before[0].missed
    tr.coast(); tr.coast()                      # coasting must not age tracks
    assert tr.coast()[0].missed == missed_before


def test_scene_change_fraction():
    from core.pipeline import scene_change_fraction
    a = np.zeros((90, 160), np.uint8)
    b = a.copy()
    assert scene_change_fraction(a, b) == 0.0            # static scene
    b[:9, :] = 255                                        # 10% of pixels change
    frac = scene_change_fraction(a, b)
    assert 0.08 < frac < 0.12
    assert scene_change_fraction(None, b) == 1.0          # first frame counts


def test_probe_prefers_upscale_when_native_face_tiny():
    from core.pipeline import probe_faces
    from models.face_detector import Face

    class StubDet:
        def detect(self, img, offset=(0, 0)):
            if img.shape[0] < 200:      # native small ROI: tiny 30px face
                return [Face(box=(5, 5, 35, 35),
                             landmarks=np.zeros((5, 2), np.float32), score=0.6)]
            return [Face(box=(10, 10, 70, 70),   # upscaled: healthy 60px face
                         landmarks=np.zeros((5, 2), np.float32), score=0.8)]

    frame = np.zeros((1080, 1920, 3), np.uint8)
    f, src, upscaled = probe_faces(StubDet(), frame, (100, 100, 180, 260))
    assert upscaled is True
    assert (f.box[2] - f.box[0]) == 60          # took the upscaled detection


def test_evaluate_logic(tmp_path, monkeypatch):
    """Ground-truth scoring: found / review / missed / false-mark buckets."""
    monkeypatch.setattr(settings, "DB_PATH", str(tmp_path / "e.db"))
    from database import db_setup, db_manager
    db_setup.init_db()
    for emp, name in [("01", "A"), ("02", "B"), ("03", "C"), ("04", "D")]:
        db_manager.add_employee(emp, name)
    r1 = fuse_day("2026-07-12", [_ev("01", "v1")])["01"]           # PRESENT
    db_manager.save_daily_record(r1)
    r2 = fuse_day("2026-07-12", [_ev("02", "v1", status=ie.CANDIDATE)])["02"]
    db_manager.save_daily_record(r2)                                # REVIEW
    r3 = fuse_day("2026-07-12", [_ev("04", "v1")])["04"]            # FALSE mark
    db_manager.save_daily_record(r3)
    rows = {r["employee_id"]: r["status"]
            for r in db_manager.get_daily_attendance("2026-07-12")}
    truth = {"01", "02", "03"}
    found = [e for e in truth if rows.get(e) == "PRESENT"]
    review = [e for e in truth if rows.get(e) == "REVIEW"]
    missed = [e for e in truth if rows.get(e, "ABSENT") == "ABSENT"]
    false_mark = [e for e, st in rows.items() if st == "PRESENT" and e not in truth]
    assert found == ["01"] and review == ["02"] and missed == ["03"]
    assert false_mark == ["04"]


# ── Phase 3e: pose-guided recognition (keypoints steer, never identify) ────
def _fake_pose_pred(n=2):
    """(56, N) pose tensor: cxcywh + conf + 17*(x,y,conf), letterbox r=1,d=0."""
    pred = np.zeros((56, n), np.float32)
    pred[0] = [200, 500][:n]; pred[1] = [300, 300][:n]
    pred[2] = 80; pred[3] = 200; pred[4] = [0.9, 0.15][:n]
    pred[5] = pred[0]; pred[6] = pred[1] - 80; pred[7] = 0.9   # nose visible
    return pred


def test_decode_pose_head_yields_boxes_and_keypoints():
    from models.person_detector_yolo import decode_yolo
    boxes, scores, kpts = decode_yolo(_fake_pose_pred(), r=1.0, dx=0, dy=0,
                                      W=1920, H=1080, floor=0.3)
    assert len(scores) == 1                       # 0.15 candidate filtered
    assert kpts is not None and kpts.shape == (1, 17, 3)
    assert abs(kpts[0, 0, 0] - 200) < 1e-3        # nose x in frame coords
    x1, y1, x2, y2 = boxes[0]
    assert abs(x1 - 160) < 1 and abs(x2 - 240) < 1


def test_decode_detection_head_still_works():
    from models.person_detector_yolo import decode_yolo
    pred = np.zeros((84, 1), np.float32)
    pred[0] = 200; pred[1] = 300; pred[2] = 80; pred[3] = 200
    pred[4] = 0.8                                 # person class score
    boxes, scores, kpts = decode_yolo(pred, 1.0, 0, 0, 1920, 1080, 0.3)
    assert len(scores) == 1 and kpts is None


def test_head_region_facing_and_back_turned():
    from core.pipeline import head_region_from_kpts
    k = np.zeros((17, 3), np.float32)
    person = (100, 100, 260, 480)
    facing, hb = head_region_from_kpts(k, person, 0.3)
    assert facing is False and hb is None         # nothing visible: back turned
    k[0] = [180, 140, 0.9]; k[1] = [170, 130, 0.8]        # nose + left eye
    k[5] = [140, 200, 0.9]; k[6] = [220, 200, 0.9]        # shoulders
    facing, hb = head_region_from_kpts(k, person, 0.3)
    assert facing is True and hb is not None
    x1, y1, x2, y2 = hb
    assert x1 >= 100 and x2 <= 260 and y1 >= 100          # clipped to person
    assert x1 <= 180 <= x2 and y1 <= 140 <= y2            # contains the nose
    assert head_region_from_kpts(None, person, 0.3) == (True, None)


# ── 13 Jul zero-detections incident: real-shaped tensors must decode ────────
def test_decode_real_shaped_pose_and_detection_tensors():
    """(C, N) with N in the thousands — the anchor dim must never be mistaken
    for the channel dim (the c>=84 rule bug)."""
    from models.person_detector_yolo import decode_yolo
    pose = np.zeros((56, 18900), np.float32)
    pose[0, 5] = 480; pose[1, 5] = 480; pose[2, 5] = 80; pose[3, 5] = 200
    pose[4, 5] = 0.88
    boxes, scores, kpts = decode_yolo(pose, 0.5, 0, 210, 1920, 1080, 0.18)
    assert len(scores) == 1 and kpts is not None

    det = np.zeros((84, 8400), np.float32)
    det[0, 7] = 480; det[1, 7] = 480; det[2, 7] = 80; det[3, 7] = 200
    det[4, 7] = 0.85
    boxes, scores, kpts = decode_yolo(det, 0.5, 0, 210, 1920, 1080, 0.18)
    assert len(scores) == 1 and kpts is None


def test_critical_settings_exist():
    """The 3a cleanup could not see getattr("NAME") references; these five
    must never silently disappear again."""
    for n in ("PERSON_DETECTION_FLOOR", "PARALLEL_VIDEOS",
              "RECOGNITION_FOCUS_LIMIT", "RECOGNITION_FOCUS_PATIENCE",
              "RECOGNITION_FOCUS_COOLDOWN"):
        assert hasattr(settings, n), n


# ── door line: crossings count only AT the doorway ──────────────────────────
def _door_line_setup():
    """A diagonal door line plus the geometry helpers to walk across it."""
    from core.pipeline import VideoPipeline, VideoJob
    W, H = 1920, 1080
    line = (0.15 * W, 0.62 * H, 0.36 * W, 0.39 * H)
    pipe = object.__new__(VideoPipeline)        # no models needed for geometry
    job = VideoJob(path="t", date="2026-07-31", start_time="10:00:00",
                   camera_id="CAM-5", camera_location="office",
                   camera_type="entrance", out_is_down=False)

    class Ident:
        identity = "08"
        def confidence(self):
            return 0.9

    return pipe, job, line, Ident()


def _cross_at(pipe, job, line, ident, t_along):
    """Walk perpendicularly through the line at parameter t along it."""
    dx, dy = line[2] - line[0], line[3] - line[1]
    n = (dx * dx + dy * dy) ** 0.5
    nx, ny = -dy / n, dx / n
    px, py = line[0] + t_along * dx, line[1] + t_along * dy
    a = (px - nx * 120, py - ny * 120, 1.0)
    b = (px + nx * 120, py + ny * 120, 2.0)
    return pipe._check_crossing([a, b], line, ident, job)


def test_door_crossing_fires_at_the_doorway():
    pipe, job, line, ident = _door_line_setup()
    assert _cross_at(pipe, job, line, ident, 0.5) is not None
    assert _cross_at(pipe, job, line, ident, 0.05) is not None


def test_door_crossing_ignores_the_lines_extension():
    """Regression, 31 Jul 2026: an employee still sitting in the room raised an
    exit alert because the crossing test used the INFINITE line through the two
    door points instead of the segment between them."""
    pipe, job, line, ident = _door_line_setup()
    for t in (2.2, -1.5, 1.4):
        assert _cross_at(pipe, job, line, ident, t) is None, \
            f"movement at t={t} is nowhere near the door and must not count"


def test_door_crossing_both_directions_differ():
    pipe, job, line, ident = _door_line_setup()
    dx, dy = line[2] - line[0], line[3] - line[1]
    n = (dx * dx + dy * dy) ** 0.5
    nx, ny = -dy / n, dx / n
    mx, my = (line[0] + line[2]) / 2, (line[1] + line[3]) / 2
    pos = (mx + nx * 120, my + ny * 120, 1.0)
    neg = (mx - nx * 120, my - ny * 120, 2.0)
    a = pipe._check_crossing([neg, pos], line, ident, job)
    b = pipe._check_crossing([pos, neg], line, ident, job)
    assert a and b and a.event_type != b.event_type


def test_door_matches_site_photo():
    """The configured door line must behave the way the site photo says.

    Photo of 31 Jul 2026: green line across the doorway, red IN arrow pointing
    down-right INTO the room, red OUT arrow pointing up-left towards the glass
    door. Walking towards the room must read IN; towards the door, OUT.
    """
    from core.pipeline import VideoPipeline, VideoJob
    from core import door

    cfg = door.load_config()
    W, H = 1920, 1080
    x1, y1, x2, y2 = cfg["line_points"]
    line = (x1 * W, y1 * H, x2 * W, y2 * H)
    pipe = object.__new__(VideoPipeline)
    job = VideoJob(path="t", date="2026-07-31", start_time="10:00:00",
                   camera_id="CAM-5", camera_location="office",
                   camera_type="entrance",
                   out_is_down=(cfg["out_direction"] == "down"))

    class Ident:
        identity = "08"
        def confidence(self):
            return 0.9

    dx, dy = line[2] - line[0], line[3] - line[1]
    n = (dx * dx + dy * dy) ** 0.5
    # unit normal (-dy, dx)/n points DOWN-RIGHT for this line = into the room
    nx, ny = -dy / n, dx / n
    assert nx > 0 and ny > 0, "normal should point into the room (down-right)"
    mx, my = (line[0] + line[2]) / 2, (line[1] + line[3]) / 2
    room = (mx + nx * 120, my + ny * 120)     # inside the office
    doorway = (mx - nx * 120, my - ny * 120)  # towards the glass door

    walking_in = pipe._check_crossing(
        [(doorway[0], doorway[1], 1.0), (room[0], room[1], 2.0)], line, Ident(), job)
    walking_out = pipe._check_crossing(
        [(room[0], room[1], 1.0), (doorway[0], doorway[1], 2.0)], line, Ident(), job)

    assert walking_in and walking_in.event_type == "IN", "towards the room = IN"
    assert walking_out and walking_out.event_type == "OUT", "towards the door = OUT"


def test_only_the_main_door_camera_raises_exit_alerts():
    """A crossing on any other camera must never announce that someone left."""
    from core import door
    w = door.DoorWatch(day="2026-07-31")
    present = w.face_confirmed_today()
    if not present:
        pytest.skip("no one marked present today to test with")
    eid = next(iter(present))
    assert w.record_exit(eid, "face", 0.9, camera_id="CAM-3") is None
    assert w.record_exit(eid, "face", 0.9, camera_id="CAM-7") is None


# ── training-image harvest: purity is the whole point ───────────────────────
def _harvest_job(tmp_path):
    from core import harvest
    return harvest.HarvestJob(videos=[], target_per_employee=500,
                              out_root=str(tmp_path))


def _hsample(**kw):
    s = dict(track_id=1, track_identity="03", track_status=ie.LOCKED,
             employee_id="03", similarity=0.72, margin=0.20, quality=0.80,
             accepted=True, ambiguous=False)
    s.update(kw)
    return s


def test_harvest_saves_only_locked_agreeing_confident_samples(tmp_path):
    j = _harvest_job(tmp_path)
    assert j._accept(_hsample()) == "03"                      # the clean case


def test_harvest_refuses_anything_doubtful(tmp_path):
    """Every one of these is a way a wrong face reached an employee's folder."""
    j = _harvest_job(tmp_path)
    for bad, why in [
        (dict(track_status=ie.CANDIDATE), "track not locked"),
        (dict(employee_id="17"), "this frame disagrees with the track"),
        (dict(ambiguous=True), "ambiguous match"),
        (dict(accepted=False), "gallery did not accept"),
        (dict(similarity=0.50), "similarity too low"),
        (dict(margin=0.05), "too close to the runner-up"),
        (dict(quality=0.20), "poor quality crop"),
        (dict(track_identity=None), "no identity at all"),
    ]:
        assert j._accept(_hsample(**bad)) is None, why


def test_harvest_body_crop_rejects_a_second_person(tmp_path):
    """A body crop with somebody else standing in it is not trainable."""
    j = _harvest_job(tmp_path)
    frame = np.zeros((720, 1280, 3), dtype=np.uint8)
    s = _hsample(frame=frame, person_box=(100, 100, 300, 600),
                 other_boxes=((1, (100, 100, 300, 600)),
                              (2, (250, 120, 450, 620))))   # overlapping
    assert j._save_body("03", s) is False


def test_harvest_dedups_near_identical_frames(tmp_path):
    j = _harvest_job(tmp_path)
    rec = j._rec("03")
    v = np.zeros(settings.FACE_EMBED_DIM, dtype=np.float32); v[0] = 1.0
    rec.embeddings.append(v)
    assert j._is_duplicate(rec, v)                    # same frame again
    w = np.zeros(settings.FACE_EMBED_DIM, dtype=np.float32); w[1] = 1.0
    assert not j._is_duplicate(rec, w)                # a genuinely new pose


# ── unregistered people get their own folders ───────────────────────────────
def _unk_sample(**kw):
    d = dict(track_id=1, track_identity=None, track_status=None,
             employee_id=None, similarity=0.12, margin=0.0, quality=0.80,
             accepted=False, ambiguous=False,
             embedding=np.zeros(settings.FACE_EMBED_DIM, dtype=np.float32))
    d.update(kw)
    return d


def test_stranger_gate(tmp_path):
    from core import harvest
    j = harvest.HarvestJob(videos=[], out_root=str(tmp_path))
    assert j._accept_unknown(_unk_sample()) is True
    # a known employee must never be filed as a new person
    assert not j._accept_unknown(_unk_sample(track_identity="03",
                                             track_status=ie.LOCKED))
    assert not j._accept_unknown(_unk_sample(accepted=True))
    assert not j._accept_unknown(_unk_sample(ambiguous=True))
    assert not j._accept_unknown(_unk_sample(quality=0.10))
    # sitting just under the accept bar is more likely a bad angle of a known
    # face than a stranger — collecting it would split that employee in two
    assert not j._accept_unknown(_unk_sample(similarity=0.42))


def test_two_strangers_do_not_share_a_folder(tmp_path):
    from core import harvest
    j = harvest.HarvestJob(videos=[], out_root=str(tmp_path))
    j.unknown = harvest.UnknownClusters(str(tmp_path))
    a = np.zeros(settings.FACE_EMBED_DIM, dtype=np.float32); a[0] = 1.0
    b = np.zeros(settings.FACE_EMBED_DIM, dtype=np.float32); b[1] = 1.0
    for tid, vec in ((10, a), (20, b)):
        for _ in range(harvest.UNKNOWN_MIN_TRACK_SAMPLES):
            j._cluster_for(_unk_sample(track_id=tid, embedding=vec))
    assert j._track_cluster[10] != j._track_cluster[20]


def test_same_stranger_reuses_their_folder(tmp_path):
    from core import harvest
    j = harvest.HarvestJob(videos=[], out_root=str(tmp_path))
    j.unknown = harvest.UnknownClusters(str(tmp_path))
    a = np.zeros(settings.FACE_EMBED_DIM, dtype=np.float32); a[0] = 1.0
    for _ in range(harvest.UNKNOWN_MIN_TRACK_SAMPLES):
        j._cluster_for(_unk_sample(track_id=10, embedding=a))
    first = j._track_cluster[10]
    for _ in range(harvest.UNKNOWN_MIN_TRACK_SAMPLES):
        j._cluster_for(_unk_sample(track_id=30, embedding=a))
    assert j._track_cluster[30] == first

    j.unknown.save()                     # and again in a LATER sweep
    assert harvest.UnknownClusters(str(tmp_path)).assign(a) == first


def test_fleeting_detection_creates_no_folder(tmp_path):
    from core import harvest
    j = harvest.HarvestJob(videos=[], out_root=str(tmp_path))
    j.unknown = harvest.UnknownClusters(str(tmp_path))
    v = np.zeros(settings.FACE_EMBED_DIM, dtype=np.float32); v[3] = 1.0
    assert j._cluster_for(_unk_sample(track_id=99, embedding=v)) is None
    assert len(j.unknown._emb) == 0


# ── best-of-N enrollment: 300 photos in, the best 40 kept ───────────────────
def _cand(vec, quality=0.8):
    from core.enrollment import EnrollResult
    v = np.asarray(vec, dtype=np.float32)
    return EnrollResult(True, quality, "", v / (np.linalg.norm(v) or 1.0))


def _make_upload(rng, n_same=210, n_varied=90, n_other=4, dim=None):
    """A realistic upload: one long burst, a few rare angles, a stray stranger."""
    dim = dim or settings.FACE_EMBED_DIM
    unit = lambda v: v / (np.linalg.norm(v) or 1.0)
    base = unit(rng.normal(size=dim))
    poses = [unit(base + 0.55 * unit(rng.normal(size=dim))) for _ in range(6)]
    out = []
    for i in range(n_same + n_varied):
        p = poses[0] if i < n_same else poses[1 + (i % 5)]
        out.append(_cand(unit(p + 0.05 * unit(rng.normal(size=dim))),
                         float(rng.uniform(0.45, 0.95))))
    for _ in range(n_other):                       # somebody else's photos
        out.append(_cand(unit(rng.normal(size=dim)), 0.9))
    rng.shuffle(out)
    return out


def _mean_pair_sim(sel):
    M = np.stack([c.embedding for c in sel])
    S = M @ M.T
    iu = np.triu_indices(len(sel), 1)
    return float(S[iu].mean())


def test_best_of_n_beats_first_n_on_variety_and_quality():
    """Taking the first 40 of an upload keeps 40 near-identical frames."""
    from core.enrollment import select_best
    rng = np.random.default_rng(7)
    cands = _make_upload(rng)
    picked, report = select_best(cands, 40)
    first_n = cands[:40]
    assert len(picked) == 40
    assert _mean_pair_sim(picked) < _mean_pair_sim(first_n)   # wider coverage
    assert (np.mean([c.quality for c in picked])
            >= np.mean([c.quality for c in first_n]))         # and not worse


def test_best_of_n_drops_someone_elses_photos():
    from core.enrollment import select_best
    rng = np.random.default_rng(11)
    _, report = select_best(_make_upload(rng, n_other=4), 40)
    assert report["outliers"] >= 3, "a stray face must not become a template"


def test_best_of_n_keeps_everything_when_under_target():
    from core.enrollment import select_best
    rng = np.random.default_rng(3)
    cands = _make_upload(rng, n_same=6, n_varied=4, n_other=0)
    picked, _ = select_best(cands, 40)
    assert len(picked) == 10


def test_replace_templates_installs_exact_set(tmp_path, monkeypatch):
    import core.embedding_gallery as eg
    monkeypatch.setattr(eg, "FACE_EMB_DIR", str(tmp_path))
    g = eg.EmbeddingGallery()
    embs = np.eye(settings.FACE_EMBED_DIM, dtype=np.float32)[:5]
    assert g.replace_templates("E1", embs) == 5
    assert g.count("E1") == 5
    # replacing again does not append
    assert g.replace_templates("E1", embs[:2]) == 2
    assert g.count("E1") == 2


# ── "find THIS person from photos" ──────────────────────────────────────────
def _unitv(dim, i):
    v = np.zeros(dim, dtype=np.float32); v[i] = 1.0
    return v


def test_target_match_needs_a_strong_similarity(tmp_path):
    from core import harvest
    D = settings.FACE_EMBED_DIM
    j = harvest.HarvestJob(videos=[], out_root=str(tmp_path),
                           targets={"Asha": _unitv(D, 0)[None]})
    assert j._match_target(_unitv(D, 0)) == "Asha"          # the same face
    assert j._match_target(_unitv(D, 5)) is None            # a different face


def test_target_match_refuses_when_two_seeded_people_are_close(tmp_path):
    """If the probe sits between two seeded people, name neither."""
    from core import harvest
    D = settings.FACE_EMBED_DIM
    a, b = _unitv(D, 0), _unitv(D, 1)
    j = harvest.HarvestJob(videos=[], out_root=str(tmp_path),
                           targets={"A": a[None], "B": b[None]})
    mid = (a + b); mid /= np.linalg.norm(mid)               # equidistant
    assert j._match_target(mid) is None
    assert j._match_target(a) == "A"                        # unambiguous still works


def test_build_reference_drops_a_different_person(tmp_path):
    """Handing in photos of two people must not seed a mixed reference."""
    from core import enrollment
    D = settings.FACE_EMBED_DIM
    unit = lambda v: v / (np.linalg.norm(v) or 1.0)
    rng = np.random.default_rng(5)
    base = unit(rng.normal(size=D))
    cands = [_cand(unit(base + 0.08 * unit(rng.normal(size=D))), 0.8)
             for _ in range(10)]
    cands += [_cand(unit(rng.normal(size=D)), 0.9) for _ in range(2)]  # intruder
    picked, report = enrollment.select_best(cands, 12)
    assert report["outliers"] >= 2
