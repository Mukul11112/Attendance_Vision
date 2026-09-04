"""
reports/excel_report.py
Excel export for one attendance date. Guarantees exactly one row per employee:
input rows come from the roster-joined query (one row per registered employee)
and are additionally de-duplicated by employee_id here as a belt-and-braces
check.
"""
from __future__ import annotations
import os
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import settings

COLUMNS = [
    ("Employee ID", "employee_id", 12),
    ("Employee Name", "name", 22),
    ("Department", "department", 16),
    ("Date", "attendance_date", 12),
    ("Status", "status", 11),
    ("Recognition Confidence", "confidence", 20),
    ("Evidence Type", "evidence_type", 13),
    ("Face Evidence Count", "face_evidence_count", 18),
    ("Body Evidence Count", "body_evidence_count", 18),
    ("Videos Seen", "videos_seen", 34),
    ("Review Required", "review_required", 14),
]

FILL = {
    "PRESENT": PatternFill("solid", fgColor="C9F2CD"),
    "REVIEW":  PatternFill("solid", fgColor="FFE9B8"),
    "ABSENT":  PatternFill("solid", fgColor="F4CFCF"),
}


def export(date: str, attendance_rows: List[dict],
           events: Optional[List[dict]] = None) -> str:
    # belt-and-braces uniqueness: keep the first row per employee_id
    seen, rows = set(), []
    for r in attendance_rows:
        emp = r.get("employee_id")
        if emp in seen:
            continue
        seen.add(emp)
        rows.append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    t = ws.cell(row=1, column=1, value=f"Attendance Report — {date}")
    t.font = Font(bold=True, size=14)
    t.alignment = Alignment(horizontal="center")

    for c, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=c, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = width

    for r, row in enumerate(rows, start=3):
        for c, (_, key, _) in enumerate(COLUMNS, start=1):
            val = row.get(key, "")
            if key == "review_required":
                val = "YES" if val in (1, True, "1") else ""
            if key == "confidence":
                val = round(float(val or 0.0), 3)
            ws.cell(row=r, column=c, value=val)
        fill = FILL.get(str(row.get("status", "")).upper())
        if fill:
            ws.cell(row=r, column=5).fill = fill

    n_present = sum(1 for r in rows if r.get("status") == "PRESENT")
    n_review = sum(1 for r in rows if r.get("status") == "REVIEW")
    n_absent = sum(1 for r in rows if r.get("status") == "ABSENT")
    s = ws.cell(row=len(rows) + 4, column=1,
                value=f"Registered: {len(rows)}   Present: {n_present}   "
                      f"Review: {n_review}   Absent: {n_absent}")
    s.font = Font(bold=True)

    if events:
        ws2 = wb.create_sheet("Gate Events (diagnostic)")
        heads = ["employee_id", "timestamp", "event_type", "camera_id",
                 "video_name", "confidence"]
        for c, h in enumerate(heads, start=1):
            ws2.cell(row=1, column=c, value=h).font = Font(bold=True)
        for r, ev in enumerate(events, start=2):
            for c, h in enumerate(heads, start=1):
                ws2.cell(row=r, column=c, value=ev.get(h, ""))

    os.makedirs(settings.EXPORTS_DIR, exist_ok=True)
    path = os.path.join(settings.EXPORTS_DIR, f"attendance_{date}.xlsx")
    wb.save(path)
    return path
